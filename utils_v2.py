"""
utils_v2.py - Multi-format extraction and LangChain-compatible vector DB builder
Fixes: Collection existence errors, deprecation warnings, proper error handling
"""

import io, os, mimetypes, tempfile, concurrent.futures as cf
from pathlib import Path
from typing import Dict, List

import fitz
import pytesseract
from PIL import Image
import docx2txt
from openpyxl import load_workbook

from langchain_community.vectorstores import Chroma
from langchain_ollama import OllamaEmbeddings  # Updated import to fix deprecation
import chromadb

_OCR_LANG = "eng"

# --- Text extraction helpers ---

def _pdf_extract_fast(file_bytes: bytes) -> str:
    """Extract text from PDF using PyMuPDF with OCR fallback for scanned pages."""
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        texts = []
        for page in doc:
            txt = page.get_text()
            if txt.strip():
                texts.append(txt)
            else:
                # Scanned page -> OCR
                try:
                    pix = page.get_pixmap()
                    img = Image.open(io.BytesIO(pix.tobytes("png")))
                    ocr_text = pytesseract.image_to_string(img, lang=_OCR_LANG)
                    texts.append(ocr_text)
                except Exception as e:
                    print(f"OCR failed on PDF page: {e}")
                    texts.append("")
        doc.close()
        return "\n".join(texts)
    except Exception as e:
        print(f"PDF extraction error: {e}")
        return ""

def _image_extract(file_bytes: bytes) -> str:
    """Extract text from image using OCR."""
    try:
        img = Image.open(io.BytesIO(file_bytes))
        config = r'--oem 3 --psm 6'
        return pytesseract.image_to_string(img, config=config, lang=_OCR_LANG)
    except Exception as e:
        print(f"Image OCR error: {e}")
        return ""

def _docx_extract(file_bytes: bytes, filename: str) -> str:
    """Extract text from DOCX file."""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        text = docx2txt.process(tmp_path)
        os.unlink(tmp_path)
        return text
    except Exception as e:
        print(f"DOCX extraction error: {e}")
        return ""

def _xlsx_extract(file_bytes: bytes, filename: str) -> str:
    """Extract text from XLSX file."""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        wb = load_workbook(tmp_path, data_only=True)
        all_texts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = [" | ".join(str(c) if c is not None else "" for c in row)
                   for row in sheet.iter_rows(values_only=True)]
            all_texts.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
        os.unlink(tmp_path)
        return "\n\n".join(all_texts)
    except Exception as e:
        print(f"XLSX extraction error: {e}")
        return ""

def _txt_extract(file_bytes: bytes) -> str:
    """Extract text from TXT/CSV file."""
    try:
        return file_bytes.decode("utf-8", errors="ignore")
    except Exception as e:
        print(f"Text extraction error: {e}")
        return ""

# File type mapping
_EXTRACTORS = {
    ".pdf": _pdf_extract_fast,
    ".docx": _docx_extract,
    ".xlsx": _xlsx_extract,
    ".txt": _txt_extract,
    ".csv": _txt_extract,
    ".png": _image_extract,
    ".jpg": _image_extract,
    ".jpeg": _image_extract,
}

def extract_text_from_file(file_bytes: bytes, filename: str) -> str:
    """Extract text from any supported file type."""
    ext = Path(filename).suffix.lower()
    if ext in _EXTRACTORS:
        extractor = _EXTRACTORS[ext]
        if ext in [".docx", ".xlsx"]:
            return extractor(file_bytes, filename)
        else:
            return extractor(file_bytes)
    # Fallback to MIME type detection
    mtype, _ = mimetypes.guess_type(filename)
    if mtype and mtype.startswith("image/"):
        return _image_extract(file_bytes)
    raise ValueError(f"Unsupported file type: {ext}")

# --- Parallel file processing ---

def process_uploaded_files(uploaded_files) -> Dict[str, str]:
    """Process multiple uploaded files in parallel."""
    def _worker(uploaded_file):
        try:
            file_bytes = uploaded_file.read()
            text = extract_text_from_file(file_bytes, uploaded_file.name)
            return uploaded_file.name, text
        except Exception as e:
            print(f"Error processing {uploaded_file.name}: {e}")
            return uploaded_file.name, ""

    results = {}
    with cf.ThreadPoolExecutor(max_workers=4) as executor:
        futures = [executor.submit(_worker, f) for f in uploaded_files]
        for future in cf.as_completed(futures):
            filename, text = future.result()
            results[filename] = text
    return results

# --- Text chunking helper ---

def _chunk_text(text: str, max_words: int = 800, overlap: int = 100) -> List[str]:
    """Split text into overlapping chunks for better RAG context."""
    if not text.strip():
        return []
    words = text.split()
    if len(words) <= max_words:
        return [text]
    
    chunks = []
    start = 0
    while start < len(words):
        end = start + max_words
        chunk = " ".join(words[start:end])
        chunks.append(chunk)
        start = end - overlap
        if start >= len(words):
            break
    return chunks

# --- LangChain-compatible vector database creation ---

def create_vector_db(texts: List[str], collection_name: str = "documents"):
    """
    Create a LangChain-compatible Chroma vectorstore.
    Returns a vectorstore object with .as_retriever() method.
    """
    if not texts or all(not t.strip() for t in texts):
        print("No valid texts provided for vector DB creation")
        return None

    try:
        # Initialize embeddings with updated import
        embeddings = OllamaEmbeddings(model="nomic-embed-text")
        
        # Prepare documents and metadata
        all_docs = []
        all_metadatas = []
        
        for doc_idx, text in enumerate(texts):
            if not text.strip():
                continue
            chunks = _chunk_text(text)
            for chunk_idx, chunk in enumerate(chunks):
                if chunk.strip():
                    all_docs.append(chunk)
                    all_metadatas.append({
                        "source": f"document_{doc_idx}",
                        "chunk": chunk_idx,
                        "doc_id": f"doc_{doc_idx}_chunk_{chunk_idx}"
                    })
        
        if not all_docs:
            print("No valid document chunks created")
            return None
        
        # Create persistent client for better reliability
        persist_directory = ".chromadb_v2"
        client = chromadb.PersistentClient(path=persist_directory)
        
        # Handle collection creation/deletion safely
        try:
            # Try to get existing collection first
            existing_collections = client.list_collections()
            collection_exists = any(c.name == collection_name for c in existing_collections)
            
            if collection_exists:
                print(f"Deleting existing collection: {collection_name}")
                client.delete_collection(name=collection_name)
                print(f"Successfully deleted collection: {collection_name}")
        except Exception as e:
            print(f"Note: Could not delete collection (may not exist): {e}")
            # This is fine, collection may not exist
        
        # Create new vectorstore
        print(f"Creating new vectorstore with {len(all_docs)} document chunks...")
        vectorstore = Chroma.from_texts(
            texts=all_docs,
            embedding=embeddings,
            metadatas=all_metadatas,
            collection_name=collection_name,
            client=client,
        )
        
        print("✅ Vector database created successfully!")
        return vectorstore
        
    except Exception as e:
        print(f"Vector DB creation error: {e}")
        print(f"Error type: {type(e).__name__}")
        return None